import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
##This script will take all the XML files in a directory and create a new Excel workbook with a worksheet for each XML file.##
# Set the directory containing the XML files
xml_dir = '/home/danielb/Desktop/Work/Prod/Prod_NSI_03242023/'

# Get a list of all the XML files in the directory
xml_files = [os.path.join(xml_dir, f) for f in os.listdir(xml_dir) if f.endswith('.xml')]

# Set the sheet names to be the filenames without the '.xml' extension
sheet_names = [os.path.splitext(os.path.basename(f))[0] for f in xml_files]

# Create a new workbook object
wb = Workbook()

# Remove the default sheet
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# Loop through each XML file and add its data to a new worksheet in the workbook with the corresponding sheet name
for i, xml_file in enumerate(xml_files):
    # Open the XML file in read mode with UTF-8 encoding
    with open(xml_file, "r", encoding="utf-8") as f:
        # Read the entire contents of the file as a string
        xml_data = f.read()
    # Parse the XML data string into an ElementTree object
    root = ET.fromstring(xml_data)
    # Create a new worksheet with the corresponding sheet name
    ws = wb.create_sheet(title=sheet_names[i])
    header = ['key', 'name', 'version', 'vendor', 'status', 'vendor-url', 'framework-version', 'bundled']
    # Write the header row to the worksheet
    for col_index, header_title in enumerate(header):
        ws.cell(row=1, column=col_index+1, value=header_title)
    # Loop through each "plugin" element in the XML data and write its child elements to a row in the worksheet
    for row_index, plugin in enumerate(root.iter('plugin')):
        ws.cell(row=row_index+2, column=1, value=plugin.find('key').text if plugin.find('key') is not None else '')
        ws.cell(row=row_index+2, column=2, value=plugin.find('name').text if plugin.find('name') is not None else '')
        ws.cell(row=row_index+2, column=3, value=plugin.find('version').text if plugin.find('version') is not None else '')
        ws.cell(row=row_index+2, column=4, value=plugin.find('vendor').text if plugin.find('vendor') is not None else '')
        ws.cell(row=row_index+2, column=5, value=plugin.find('status').text if plugin.find('status') is not None else '')
        ws.cell(row=row_index+2, column=6, value=plugin.find('vendor-url').text if plugin.find('vendor-url') is not None else '')
        ws.cell(row=row_index+2, column=7, value=plugin.find('framework-version').text if plugin.find('framework-version') is not None else '')
        ws.cell(row=row_index+2, column=8, value=plugin.find('bundled').text if plugin.find('bundled') is not None else '')

# Save the workbook to a file
wb.save('/home/danielb/Desktop/Work/Prod/Prod_NSI_03242023/Prod_Plugins.xlsx')
