import re
import openpyxl
import glob
### This is to combine all the access logs into one excel file###
# Regex pattern to extract the specified details
pattern = (r'(?P<Origin_IP>\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}) '  # IP address
          r'(?P<Request_ID>\d+x\d+x\d+) '  # Request ID
          r'(?P<Username>\S+) \['  # Username
          r'(?P<Timestamp>.*?)\] '  # Timestamp
          r'"(?P<Method_Endpoint>.*?)" '  # Method & Endpoint
          r'(?P<HTTP_status_code>\d+) '  # HTTP status code
          r'(?P<Bytes_sent>\d+) '  # Bytes sent
          r'(?P<Time_millis>\d+) '  # Time taken in millis
          r'"(?P<Accessed_URL>.*?)" '  # Accessed URL
          r'"(?P<Client>.*?)" '  # Client utilized
          r'"(?P<Session_ID>.*?)"')  # Session ID

# List to store matched data
data = []

# Specify the path where the log files are located and the matching pattern
folder_path = "/home/sithlord/Downloads/Jira_c262175d2a93417780652ad16d4c2434_ip-10-5-96-145-us-gov-west-1-compute-internal_support_2023-09-14-13-25-42/tomcat-access-logs/"
file_pattern = "access_log.*"

# Create a new workbook
wb = openpyxl.Workbook()
# Remove the default created sheet (if you want to)
wb.remove(wb.active)

# Define a mapping from regex group names to header names
header_map = {
    'Origin_IP': 'Origin IP',
    'Request_ID': 'Request ID',
    'Username': 'Username',
    'Timestamp': 'Timestamp',
    'Method_Endpoint': 'Method & Endpoint',
    'HTTP_status_code': 'HTTP status code',
    'Bytes_sent': 'Bytes sent',
    'Time_millis': 'Time Taken to process the request in millis',
    'Accessed_URL': 'Accessed URL',
    'Client': 'Client utilized',
    'Session_ID': 'Session ID'
}

# List of desired headers
headers = ['Origin IP', 'Request ID', 'Username', 'Timestamp', 
           'Method & Endpoint', 'HTTP status code', 'Bytes sent', 
           'Time Taken to process the request in millis', 'Accessed URL', 
           'Client utilized', 'Session ID']

log_files = glob.glob(folder_path + file_pattern)

for log_file in log_files:
    data = []
    with open(log_file, 'r') as file:
        for line in file:
            match = re.match(pattern, line)
            if match:
                data.append(match.groupdict())

    mapped_data = []
    for row in data:
        mapped_data.append({header_map[key]: value for key, value in row.items()})

    # Create a new sheet for this log file
    sheet_name = log_file.split('/')[-1].split('.')[1]  # Extracting the date as sheet name
    ws = wb.create_sheet(title=sheet_name)

    # Add headers to the worksheet
    for col_num, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        ws['{}1'.format(col_letter)] = header

    # Add data to the worksheet
    for row_num, row_data in enumerate(mapped_data, 2):
        for col_num, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws['{}{}'.format(col_letter, row_num)] = row_data[header]

# Save the workbook to an xlsx file
xlsx_filename = "combined_access_log.xlsx"
wb.save(folder_path + xlsx_filename)