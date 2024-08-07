import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook

# Set the directory containing the XML files
xml_dir = '/home/danielb/Desktop/Work/Script for Work/'

# Get a list of all the XML files in the directory
xml_files = [os.path.join(xml_dir, f) for f in os.listdir(xml_dir) if f.endswith('.xml')]

# Set the sheet names to be the filenames without the '.xml' extension
sheet_names = [os.path.splitext(os.path.basename(f))[0] for f in xml_files]

# Create a new workbook object
wb = Workbook()

# Remove the default sheet
default_sheet = wb['Sheet']
wb.remove(default_sheet)

# Loop through each XML file and add its data to a new worksheet in the workbook with the corresponding sheet name
for i, xml_file in enumerate(xml_files):
    with open(xml_file, "r", encoding="utf-8") as f:
        xml_data = f.read()
    root = ET.fromstring(xml_data)
    ws = wb.create_sheet(title=sheet_names[i])

    header = ['key', 'slug', 'name', 'type', 'git-lfs']
    
    for col_index, header_title in enumerate(header):
        ws.cell(row=1, column=col_index+1, value=header_title)

    project_key = ''
    row_index = 0
    for element in root.iter():
        if element.tag == 'project':
            project_key_element = element.find('key')
            if project_key_element is not None:
                project_key = project_key_element.text
        elif element.tag == 'repository':
            row_index += 1  # only increment for 'repository' elements
            for col_index, tag in enumerate(header):
                if tag == 'key':
                    ws.cell(row=row_index+2, column=col_index+1, value=project_key)
                else:
                    tag_element = element.find(tag)
                    if tag_element is not None:
                        if tag == 'git-lfs':
                            ws.cell(row=row_index+2, column=col_index+1, value=tag_element.find('enabled').text)
                        else:
                            ws.cell(row=row_index+2, column=col_index+1, value=tag_element.text)
                    else:
                        ws.cell(row=row_index+2, column=col_index+1, value='')
# Save the workbook to a file
wb.save('/home/danielb/Desktop/Work/Script for Work/bb-lfs.xlsx')